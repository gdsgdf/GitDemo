import openpyxl
book=openpyxl.load_workbook("C:\\Users\\Dell\\Exceldemotestcase.xlsx")
sheet=book.active
Dict={}
class HomePageData:
    test_homepag_data=[{"firstname":"Laxmi","email":"Rama","gender":"Female"},{"firstname":"Bunny","email":"Karthikeya","gender":"Male"}]
    @staticmethod
    def GetTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\Dell\\Exceldemotestcase.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[(sheet.cell(row=1, column=j).value)] = (sheet.cell(row=i, column=j).value)
        return[Dict]


