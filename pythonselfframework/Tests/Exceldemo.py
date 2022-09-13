import openpyxl

book=openpyxl.load_workbook("C:\\Users\\Dell\\Exceldemotestcase.xlsx")
sheet=book.active
cell=sheet.cell(row=2,column=2)
Dict={}
print(cell)
rows=sheet.max_row
columns=sheet.max_column
print(rows)
print(columns)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="testcase2":
        for j in range(2,sheet.max_column+1):
            Dict[(sheet.cell(row=1,column=j).value)]=(sheet.cell(row=i,column=j).value)
print(Dict)


